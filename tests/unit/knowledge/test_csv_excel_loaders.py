"""Tests for CSVLoader and ExcelLoader."""

from __future__ import annotations

import tempfile
from pathlib import Path
from unittest.mock import patch

import pytest

from syrin.knowledge._document import Document


class TestCSVLoader:
    """Tests for CSVLoader."""

    @pytest.fixture
    def temp_csv(self) -> str:
        """Create a temporary CSV file."""
        fd, path = tempfile.mkstemp(suffix=".csv")
        Path(path).write_text("name,category,shares\nPromoter,A,1250000\nPublic,B,750000")
        yield path
        Path(path).unlink(missing_ok=True)

    def test_load_csv(self, temp_csv: str) -> None:
        """CSVLoader loads CSV content."""
        from syrin.knowledge.loaders import CSVLoader

        loader = CSVLoader(temp_csv)
        docs = loader.load()

        assert len(docs) >= 1
        assert all(isinstance(d, Document) for d in docs)
        assert docs[0].source == temp_csv
        assert docs[0].source_type == "csv"
        assert "Promoter" in docs[0].content or "name" in docs[0].content

    def test_file_not_found_raises(self) -> None:
        """CSVLoader raises FileNotFoundError when file does not exist."""
        from syrin.knowledge.loaders import CSVLoader

        loader = CSVLoader("/nonexistent/file.csv")
        with pytest.raises(FileNotFoundError, match="does not exist"):
            loader.load()

    def test_rows_per_document(self, temp_csv: str) -> None:
        """CSVLoader respects rows_per_document when set."""
        from syrin.knowledge.loaders import CSVLoader

        loader = CSVLoader(temp_csv, rows_per_document=1)
        docs = loader.load()

        assert len(docs) >= 1
        assert all(d.source_type == "csv" for d in docs)

    def test_accepts_path(self, temp_csv: str) -> None:
        """CSVLoader accepts Path object."""
        from syrin.knowledge.loaders import CSVLoader

        loader = CSVLoader(Path(temp_csv))
        docs = loader.load()
        assert docs[0].source == temp_csv

    @pytest.mark.asyncio
    async def test_aload(self, temp_csv: str) -> None:
        """CSVLoader aload() works."""
        from syrin.knowledge.loaders import CSVLoader

        loader = CSVLoader(temp_csv)
        docs = await loader.aload()
        assert len(docs) >= 1


class TestExcelLoader:
    """Tests for ExcelLoader."""

    @pytest.fixture
    def temp_xlsx(self) -> str:
        """Create a temporary XLSX file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl required for Excel tests")
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        Path(path).unlink(missing_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws:
            ws["A1"] = "Category"
            ws["B1"] = "Shares"
            ws["A2"] = "Promoter"
            ws["B2"] = 1250000
        wb.save(path)
        yield path
        Path(path).unlink(missing_ok=True)

    def test_load_excel(self, temp_xlsx: str) -> None:
        """ExcelLoader loads Excel content."""
        from syrin.knowledge.loaders import ExcelLoader

        loader = ExcelLoader(temp_xlsx)
        docs = loader.load()

        assert len(docs) >= 1
        assert docs[0].source == temp_xlsx
        assert docs[0].source_type == "xlsx"
        assert "Promoter" in docs[0].content or "Category" in docs[0].content

    def test_file_not_found_raises(self) -> None:
        """ExcelLoader raises FileNotFoundError when file does not exist."""
        from syrin.knowledge.loaders import ExcelLoader

        loader = ExcelLoader("/nonexistent/file.xlsx")
        with pytest.raises(FileNotFoundError, match="does not exist"):
            loader.load()

    def test_openpyxl_not_installed_raises(self) -> None:
        """ExcelLoader raises ImportError when openpyxl not installed."""
        from syrin.knowledge.loaders import ExcelLoader

        with patch("syrin.knowledge.loaders._excel._check_openpyxl") as mock_check:
            mock_check.side_effect = ImportError(
                "openpyxl is required. Install with: pip install syrin[excel]"
            )
            loader = ExcelLoader("/any/file.xlsx")
            with pytest.raises(ImportError, match="openpyxl"):
                loader.load()
