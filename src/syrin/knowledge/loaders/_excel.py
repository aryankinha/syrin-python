"""Excel loader using openpyxl."""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Protocol

from syrin.knowledge._document import Document, DocumentMetadata


class _ExcelSheetLike(Protocol):
    """Protocol for openpyxl Worksheet (iter_rows)."""

    def iter_rows(self, values_only: bool = False) -> Iterable[tuple[object, ...]]:
        """Iterate rows. values_only=True yields tuples of cell values."""
        ...


def _check_openpyxl() -> None:
    """Ensure openpyxl is installed."""
    try:
        import openpyxl  # noqa: F401
    except ImportError as err:
        raise ImportError(
            "openpyxl is required for ExcelLoader. Install with: pip install syrin[excel]"
        ) from err


def _sheet_to_text(sheet: _ExcelSheetLike) -> str:
    """Convert openpyxl worksheet to text (markdown-style)."""
    data: list[list[str]] = []
    for row_tuple in sheet.iter_rows(values_only=True):
        str_row = [str(c) if c is not None else "" for c in row_tuple]
        data.append(str_row)

    if not data:
        return ""

    lines: list[str] = []
    for i, row_vals in enumerate(data):
        lines.append(" | ".join(row_vals))
        if i == 0:
            lines.append("---" * (len(row_vals) or 1))
    return "\n".join(lines)


class ExcelLoader:
    """Load Excel files (XLSX). Each sheet becomes one or more Documents.

    Requires: pip install syrin[excel]
    """

    def __init__(
        self,
        path: str | Path,
        *,
        sheets: list[str] | None = None,
    ) -> None:
        """Initialize ExcelLoader.

        Args:
            path: Path to the Excel file.
            sheets: Sheet names to load. None = all sheets.
        """
        self._path = Path(path)
        self._sheets = sheets

    @property
    def path(self) -> Path:
        """Path to the Excel file."""
        return self._path

    def load(self) -> list[Document]:
        """Load the Excel file.

        Returns:
            List of Documents, one per sheet (or per selected sheet).
        """
        _check_openpyxl()

        if not self._path.exists():
            raise FileNotFoundError(f"Excel file does not exist: {self._path}")

        import openpyxl

        wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        docs: list[Document] = []

        sheet_names = self._sheets if self._sheets is not None else wb.sheetnames
        for name in sheet_names:
            if name not in wb.sheetnames:
                continue
            sheet = wb[name]
            text = _sheet_to_text(sheet)
            meta: DocumentMetadata = {
                "sheet": name,
            }
            docs.append(
                Document(
                    content=text,
                    source=str(self._path),
                    source_type="xlsx",
                    metadata=meta,
                )
            )

        wb.close()

        if not docs:
            docs.append(
                Document(
                    content="",
                    source=str(self._path),
                    source_type="xlsx",
                    metadata={},
                )
            )

        return docs

    async def aload(self) -> list[Document]:
        """Load Excel asynchronously."""
        import asyncio

        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, self.load)


__all__ = ["ExcelLoader"]
